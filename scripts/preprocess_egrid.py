#!/usr/bin/env python3
"""
Preprocess EPA eGRID Excel → data/egrid_subregions.json

REAL DATA SOURCE:
  https://www.epa.gov/egrid/download-data
  Download: eGRID2023_data.xlsx  (or latest year)

KEY SHEET: "SRL23" (Subregion-level data for 2023)
  or "SUBRGN" in older versions

KEY COLUMNS:
  SUBRGN   - subregion code (e.g. CAMX, ERCT, RFCE)
  SRNAME   - subregion name
  SRLCO2RTA - CO2 output emission rate (lb/MWh) — this is what we want

USAGE:
  pip install openpyxl
  python3 scripts/preprocess_egrid.py
  python3 scripts/preprocess_egrid.py --input path/to/eGRID2023_data.xlsx
"""

import json
import argparse
import os
import sys

def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument('--input',  default='scripts/raw/eGRID2023_data.xlsx')
    p.add_argument('--output', default='data/egrid_subregions.json')
    return p.parse_args()

# Approximate centroids per subregion (for map display)
CENTROIDS = {
    'AZNM': {'lat':34.5,'lng':-112.0},
    'CAMX': {'lat':36.8,'lng':-119.5},
    'ERCT': {'lat':31.5,'lng':-99.0},
    'FRCC': {'lat':28.0,'lng':-82.0},
    'MROE': {'lat':44.0,'lng':-89.0},
    'MROW': {'lat':44.5,'lng':-96.0},
    'NEWE': {'lat':43.5,'lng':-71.5},
    'NWPP': {'lat':46.0,'lng':-117.0},
    'NYCW': {'lat':40.7,'lng':-74.0},
    'NYLI': {'lat':40.8,'lng':-73.2},
    'NYUP': {'lat':43.0,'lng':-76.0},
    'RFCE': {'lat':39.5,'lng':-76.0},
    'RFCM': {'lat':43.5,'lng':-84.5},
    'RFCW': {'lat':40.1,'lng':-82.5},
    'RMPA': {'lat':39.5,'lng':-105.0},
    'SPNO': {'lat':39.0,'lng':-97.5},
    'SPSO': {'lat':35.5,'lng':-97.0},
    'SRMV': {'lat':32.5,'lng':-91.5},
    'SRMW': {'lat':38.5,'lng':-90.0},
    'SRSO': {'lat':32.5,'lng':-86.5},
    'SRTV': {'lat':35.5,'lng':-86.5},
    'SRVC': {'lat':35.5,'lng':-79.5},
    'AKGD': {'lat':61.2,'lng':-149.9},
    'AKMS': {'lat':64.8,'lng':-147.7},
    'HIOA': {'lat':21.3,'lng':-157.8},
    'HIMS': {'lat':20.5,'lng':-156.3},
}

# States per subregion (approximate)
STATES = {
    'AZNM':['AZ','NM','NV'],'CAMX':['CA'],'ERCT':['TX'],'FRCC':['FL'],
    'MROE':['WI'],'MROW':['MN','ND','SD','NE','IA'],
    'NEWE':['ME','NH','VT','MA','RI','CT'],'NWPP':['WA','OR','ID','MT'],
    'NYCW':['NY'],'NYLI':['NY'],'NYUP':['NY'],
    'RFCE':['NJ','PA','MD','DE','DC'],'RFCM':['MI'],'RFCW':['OH','IN','KY','WV'],
    'RMPA':['CO','WY'],'SPNO':['KS','MO'],'SPSO':['OK','AR'],
    'SRMV':['LA','MS'],'SRMW':['IL'],'SRSO':['AL','GA','FL partial'],
    'SRTV':['TN'],'SRVC':['VA','NC','SC'],
    'AKGD':['AK'],'AKMS':['AK'],'HIOA':['HI'],'HIMS':['HI'],
}

def process_egrid(input_path, output_path):
    if not os.path.exists(input_path):
        print(f"ERROR: Input file not found: {input_path}")
        print("Download from: https://www.epa.gov/egrid/download-data")
        sys.exit(1)

    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    # Find the subregion sheet (varies by year)
    sheet = None
    for name in wb.sheetnames:
        if 'SRL' in name.upper() or 'SUBR' in name.upper():
            sheet = wb[name]
            print(f"Using sheet: {name}")
            break

    if not sheet:
        print(f"ERROR: Could not find subregion sheet. Available: {wb.sheetnames}")
        sys.exit(1)

    rows = list(sheet.iter_rows(values_only=True))
    # Find header row (contains SUBRGN)
    header_row = None
    for i, row in enumerate(rows[:10]):
        if row and any(str(c or '').upper() == 'SUBRGN' for c in row):
            header_row = i
            break

    if header_row is None:
        print(f"ERROR: Could not find header row with SUBRGN column")
        print(f"First 5 rows: {rows[:5]}")
        sys.exit(1)

    headers = [str(c or '').strip().upper() for c in rows[header_row]]

    def col(*names):
        for name in names:
            for h in headers:
                if h == name or h.startswith(name):
                    return headers.index(h)
        return None

    idx_code = col('SUBRGN')
    idx_name = col('SRNAME')
    # eGRID 2023 uses SRCO2RTA (subregion CO2 output rate, lb/MWh)
    idx_co2  = col('SRCO2RTA', 'SRLCO2RTA', 'SRLCO2RT', 'SRCO2RT')

    if idx_co2 is None:
        # Try other common names
        for candidate in ['SRLCO2','CO2RTA','CO2RTAN','LBSCO2']:
            idx_co2 = col(candidate)
            if idx_co2 is not None:
                break

    print(f"Column indices: code={idx_code}, name={idx_name}, co2={idx_co2}")
    if idx_co2 is None:
        print(f"WARNING: CO2 column not found. Headers: {headers}")

    subregions = []
    for row in rows[header_row+1:]:
        if not row or row[idx_code] is None:
            continue
        code = str(row[idx_code]).strip().upper()
        if not code or code == 'SUBRGN':
            continue

        try:
            co2 = float(row[idx_co2]) if idx_co2 is not None and row[idx_co2] else None
        except (ValueError, TypeError):
            co2 = None

        name = str(row[idx_name]).strip() if idx_name is not None and row[idx_name] else code

        subregions.append({
            'code':           code,
            'name':           name,
            'co2_lbs_per_mwh': round(co2, 1) if co2 else None,
            'centroid':       CENTROIDS.get(code, {'lat':39,'lng':-98}),
            'states':         STATES.get(code, []),
        })

    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    with open(output_path, 'w') as f:
        json.dump({'subregions': subregions}, f, indent=2)

    print(f"✅ Wrote {len(subregions)} subregions → {output_path}")
    for s in subregions:
        print(f"   {s['code']}: {s['co2_lbs_per_mwh']} lb CO₂/MWh")

if __name__ == '__main__':
    args = parse_args()
    process_egrid(args.input, args.output)
